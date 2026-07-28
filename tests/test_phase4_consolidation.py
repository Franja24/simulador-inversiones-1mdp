"""Consolidación financiera y temporal de backtesting/walk-forward."""

from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.orm import Session

from config.quant_score import BacktestConfig, QuantScoreConfig
from domain.quant import BacktestTrade
from repositories.market_history_repository import MarketHistoryRepository
from services.allocation_service import RestrictedEqualWeightAllocator
from services.backtest_service import BacktestService
from services.quant_report_service import QuantReportService
from services.quant_score_service import QuantScoreService
from tests.test_phase4_quant import seed_quant_history


def test_restricted_equal_weight_and_cash_policy() -> None:
    allocator = RestrictedEqualWeightAllocator()
    full = allocator.allocate(
        ["A", "B", "C", "D", "E"],
        top_n=5,
        maximum_symbol_weight=0.20,
        allow_cash=False,
    )
    assert full.cash_weight == 0
    assert set(full.weights.values()) == {0.20}
    cash = allocator.allocate(
        ["A", "B", "C"],
        top_n=5,
        maximum_symbol_weight=0.25,
        allow_cash=True,
    )
    assert sum(cash.weights.values()) == pytest.approx(0.75)
    assert cash.cash_weight == pytest.approx(0.25)
    assert cash.warnings
    with pytest.raises(ValueError, match="maximum_symbol_weight"):
        allocator.allocate(
            ["A", "B", "C"],
            top_n=5,
            maximum_symbol_weight=0.25,
            allow_cash=False,
        )
    empty = allocator.allocate(
        [], top_n=5, maximum_symbol_weight=0.20, allow_cash=True
    )
    assert empty.cash_weight == 1
    with pytest.raises(ValueError, match="elegibles"):
        allocator.allocate(
            [], top_n=5, maximum_symbol_weight=0.20, allow_cash=False
        )


def test_backtest_config_blocks_overlap_and_impossible_allocation() -> None:
    BacktestConfig(
        rebalance_frequency=5,
        holding_period=5,
        top_n=5,
        maximum_symbol_weight=0.20,
        allow_cash=False,
    )
    with pytest.raises(ValidationError, match="posiciones solapadas"):
        BacktestConfig(rebalance_frequency=4, holding_period=5)
    with pytest.raises(ValidationError, match="asignar 100"):
        BacktestConfig(
            top_n=3,
            maximum_symbol_weight=0.25,
            allow_cash=False,
        )


def test_execution_prices_respect_d_plus_one_missing_days_and_limit(
    session: Session,
) -> None:
    data = seed_quant_history(session, 30)
    service = BacktestService(session)
    rows = data["ALFA.MX"]
    execution = rows[5].date
    requested_exit = rows[10].date
    prices = service._execution_prices(
        "ALFA.MX", execution, requested_exit, requested_exit
    )
    assert prices is not None
    assert prices[2] == execution
    assert prices[3] == requested_exit

    missing_entry = rows[6]
    session.delete(
        next(
            item
            for item in MarketHistoryRepository(session).list_history("ALFA.MX")
            if item.date == missing_entry.date
        )
    )
    session.commit()
    delayed = service._execution_prices(
        "ALFA.MX", missing_entry.date, rows[10].date, rows[10].date
    )
    assert delayed is not None
    assert delayed[2] == rows[7].date
    assert service._execution_prices(
        "ALFA.MX", rows[5].date, rows[11].date, rows[10].date
    ) is None
    assert service._execution_prices(
        "NONE", rows[5].date, rows[10].date, rows[10].date
    ) is None


def test_calculate_symbol_uses_transversal_universe_sizes(session: Session) -> None:
    data = seed_quant_history(session, 70)
    effective = data["ALFA.MX"][-1].date
    service = QuantScoreService(session)
    one = service.calculate_symbol(
        "ALFA.MX", effective, "^MXX", universe=["ALFA.MX"]
    )
    two = service.calculate_symbol(
        "ALFA.MX", effective, "^MXX", universe=["ALFA.MX", "BETA.MX"]
    )
    three = service.calculate_symbol(
        "ALFA.MX",
        effective,
        "^MXX",
        universe=["ALFA.MX", "BETA.MX", "GAMA.MX"],
    )
    four = service.calculate_symbol(
        "ALFA.MX",
        effective,
        "^MXX",
        universe=["ALFA.MX", "BETA.MX", "GAMA.MX", "SIN_DATOS.MX"],
    )
    assert one.data_status == two.data_status == "LIMITED_UNIVERSE"
    assert one.confidence < two.confidence < three.confidence
    assert three.data_status == "OK"
    assert four.data_status == "OK"


def test_calculate_symbol_requires_explicit_or_persisted_universe(
    session: Session,
) -> None:
    with pytest.raises(ValueError, match="universo explícito"):
        QuantScoreService(session).calculate_symbol(
            "ALFA.MX", date(2025, 1, 2), "^MXX"
        )


def test_common_valuation_applies_costs_and_constraints_to_comparators(
    session: Session,
) -> None:
    data = seed_quant_history(session, 75)
    start = data["ALFA.MX"][45].date
    end = data["ALFA.MX"][-1].date
    score = QuantScoreConfig(minimum_history_rows=20)
    base = BacktestConfig(
        top_n=2,
        rebalance_frequency=5,
        holding_period=5,
        maximum_symbol_weight=0.5,
        transaction_cost_bps=0,
    )
    free = BacktestService(session).run(
        ["ALFA.MX", "BETA.MX", "GAMA.MX"], start, end, "^MXX", score, base
    )
    expensive = BacktestService(session).run(
        ["ALFA.MX", "BETA.MX", "GAMA.MX"],
        start,
        end,
        "^MXX",
        score,
        base.model_copy(update={"transaction_cost_bps": 10_000}),
    )
    for strategy in [
        "aqs", "equal_weight_universe", "random_selection", "momentum_20"
    ]:
        assert expensive.comparison[strategy] < free.comparison[strategy]
    free_winner = max(free.comparison, key=free.comparison.get)  # type: ignore[arg-type]
    costly_winner = max(  # type: ignore[arg-type]
        expensive.comparison, key=expensive.comparison.get
    )
    assert costly_winner in {"cash", "benchmark"}
    assert costly_winner != free_winner


def test_manual_curve_metrics_drawdown_turnover_and_profit_factor() -> None:
    trades = [
        BacktestTrade(
            signal_date=date(2025, 1, 1),
            execution_date=date(2025, 1, 2),
            exit_date=date(2025, 1, 3),
            symbol="A",
            entry_price=100,
            exit_price=110,
            gross_return=0.10,
            net_return=0.10,
            weight=1,
            transaction_cost=0,
        ),
        BacktestTrade(
            signal_date=date(2025, 1, 3),
            execution_date=date(2025, 1, 4),
            exit_date=date(2025, 1, 5),
            symbol="B",
            entry_price=100,
            exit_price=80,
            gross_return=-0.20,
            net_return=-0.20,
            weight=1,
            transaction_cost=0,
        ),
    ]
    metrics = BacktestService._metrics(
        pd.Series([0.10, -0.20, 0.10]),
        pd.Series([0.05, -0.10, 0.02]),
        trades,
        turnover=1.5,
    )
    assert metrics.cumulative_return == pytest.approx(-0.032)
    assert metrics.max_drawdown == pytest.approx(-0.20)
    assert metrics.turnover == 1.5
    assert metrics.profit_factor == pytest.approx(0.5)
    assert BacktestService._turnover(
        {"A": 0.5}, 0.5, {"B": 0.5}, 0.5
    ) == pytest.approx(0.5)


def test_walk_forward_real_oos_reproducible_and_future_safe(
    session: Session, tmp_path: Path,
) -> None:
    data = seed_quant_history(session, 90)
    start = data["ALFA.MX"][0].date
    end = data["ALFA.MX"][-1].date
    score = QuantScoreConfig(minimum_history_rows=20)
    config = BacktestConfig(
        top_n=2,
        rebalance_frequency=5,
        holding_period=5,
        maximum_symbol_weight=0.5,
        calibration_sessions=20,
        evaluation_sessions=15,
    )
    service = BacktestService(session)
    first = service.run_walk_forward(
        ["ALFA.MX", "BETA.MX", "GAMA.MX"], start, end, "^MXX", score, config
    )
    second = service.run_walk_forward(
        ["ALFA.MX", "BETA.MX", "GAMA.MX"], start, end, "^MXX", score, config
    )
    assert first.run_id == second.run_id
    assert first.metrics == second.metrics
    assert first.windows
    assert all(
        window.training_end < window.evaluation_start
        for window in first.windows
    )
    assert all(
        window.evaluation_start <= trade.signal_date <= window.evaluation_end
        for window in first.windows
        for trade in window.trades
    )
    first_window = first.windows[0].model_copy()
    future = data["GAMA.MX"][-1].model_copy(
        update={
            "open": 9_900,
            "high": 10_100,
            "low": 9_800,
            "close": 10_000,
            "adj_close": 10_000,
            "volume": 99_000_000,
        }
    )
    MarketHistoryRepository(session).upsert_many([future])
    session.commit()
    changed = service.run_walk_forward(
        ["ALFA.MX", "BETA.MX", "GAMA.MX"], start, end, "^MXX", score, config
    )
    assert changed.windows[0].cumulative_return == first_window.cumulative_return
    assert changed.windows[0].trades == first_window.trades
    assert changed.run_id != first.run_id
    assert first.oos_equity_curve
    assert first.metrics.aggregate_oos_return == pytest.approx(
        first.oos_equity_curve[-1]["equity"] - 1  # type: ignore[operator]
    )
    report = QuantReportService().generate(
        [], [], output_dir=tmp_path, walk_forward=first
    )
    workbook = load_workbook(report, read_only=True)
    assert {
        "Walk-forward resumen",
        "Ventanas OOS",
        "Curva OOS",
        "Operaciones OOS",
        "Comparación OOS",
    }.issubset(workbook.sheetnames)
    workbook.close()


def test_walk_forward_rejects_insufficient_window(session: Session) -> None:
    data = seed_quant_history(session, 30)
    with pytest.raises(ValueError, match="walk-forward"):
        BacktestService(session).run_walk_forward(
            ["ALFA.MX", "BETA.MX", "GAMA.MX"],
            data["ALFA.MX"][0].date,
            data["ALFA.MX"][-1].date,
            "^MXX",
            QuantScoreConfig(minimum_history_rows=20),
            BacktestConfig(calibration_sessions=30, evaluation_sessions=10),
        )
