"""Pruebas del reporte Excel y sus reglas."""

from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from config.settings import Settings
from database.models import AuditLogModel, TransactionType
from domain.portfolio import PortfolioCreate
from repositories.price_repository import SqlPriceRepository
from services.market_data_service import MarketDataService
from services.portfolio_service import PortfolioService
from services.report_service import ReportService
from services.transaction_service import TransactionService
from tests.test_transaction_service import operation


def rules(path: Path) -> dict[str, object]:
    """Lee la hoja Reglas como un diccionario."""
    workbook = load_workbook(path, data_only=True)
    result = {
        str(row[0].value): row[1].value
        for row in workbook["Reglas"].iter_rows(min_row=2)
    }
    workbook.close()
    return result


def test_empty_report_has_expected_sheets_and_rules(
    session: Session, portfolio_id: int, tmp_path: Path
) -> None:
    path = ReportService(session).generate(portfolio_id, tmp_path)
    workbook = load_workbook(path)
    assert workbook.sheetnames == ReportService.SHEETS
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert all(cell.font.bold for cell in sheet[1])
        assert sheet.auto_filter.ref is not None
    assert workbook["Reglas"]["B2"].number_format == "0.00%"
    workbook.close()
    values = rules(path)
    assert values["Cantidad actual de emisoras"] == 0
    assert values["Estado general"] == "INCUMPLE"
    assert "REPORT_GENERATED" in list(session.scalars(select(AuditLogModel.action)))


def test_rules_report_complies(
    session: Session, portfolio_id: int, tmp_path: Path
) -> None:
    TransactionService(session).register(
        operation(portfolio_id, TransactionType.BUY, "100", "10")
    )
    MarketDataService(session).save_price("AMXL.MX", Decimal("11"), datetime.now(UTC))
    settings = Settings(min_different_symbols=1, max_position_weight=0.50)
    values = rules(ReportService(session, settings).generate(portfolio_id, tmp_path))
    assert values["Cumplimiento del mínimo"] == "CUMPLE"
    assert values["Estado general"] == "CUMPLE"


def test_rules_report_detects_missing_symbol_and_price(
    session: Session, portfolio_id: int, tmp_path: Path
) -> None:
    TransactionService(session).register(
        operation(portfolio_id, TransactionType.BUY, "100", "10")
    )
    settings = Settings(min_different_symbols=2, max_position_weight=0.50)
    values = rules(ReportService(session, settings).generate(portfolio_id, tmp_path))
    assert values["Cumplimiento del mínimo"] == "INCUMPLE"
    assert values["Precios faltantes"] == "AMXL.MX"
    assert values["Estado general"] == "INCUMPLE"


def test_rules_report_detects_concentration(
    session: Session, portfolio_id: int, tmp_path: Path
) -> None:
    TransactionService(session).register(
        operation(portfolio_id, TransactionType.BUY, "60000", "10")
    )
    MarketDataService(session).save_price("AMXL.MX", Decimal("10"), datetime.now(UTC))
    settings = Settings(min_different_symbols=1, max_position_weight=0.50)
    values = rules(ReportService(session, settings).generate(portfolio_id, tmp_path))
    assert values["Posiciones sobreconcentradas"] == "AMXL.MX"
    assert values["Estado general"] == "INCUMPLE"


def test_rules_report_detects_stale_price(
    session: Session, portfolio_id: int, tmp_path: Path
) -> None:
    TransactionService(session).register(
        operation(portfolio_id, TransactionType.BUY, "100", "10")
    )
    MarketDataService(session).save_price(
        "AMXL.MX", Decimal("10"), datetime.now(UTC) - timedelta(days=7)
    )
    settings = Settings(min_different_symbols=1, max_position_weight=0.50)
    values = rules(ReportService(session, settings).generate(portfolio_id, tmp_path))
    assert values["Precios desactualizados"] == "AMXL.MX"
    assert values["Estado general"] == "INCUMPLE"


def test_report_excludes_prices_from_other_portfolios(
    session: Session, portfolio_id: int, tmp_path: Path
) -> None:
    other = PortfolioService(session).create(
        PortfolioCreate(name="Otro", initial_capital=Decimal("1000000"))
    )
    TransactionService(session).register(
        operation(portfolio_id, TransactionType.BUY, "100", "10", "AMXL.MX")
    )
    TransactionService(session).register(
        operation(other.id, TransactionType.BUY, "100", "20", "WALMEX.MX")
    )
    market = MarketDataService(session)
    market.save_price("AMXL.MX", Decimal("11"), datetime.now(UTC))
    market.save_price("WALMEX.MX", Decimal("21"), datetime.now(UTC))
    assert {
        item.symbol
        for item in SqlPriceRepository(session).list_for_symbols({"AMXL.MX"})
    } == {"AMXL.MX"}

    path = ReportService(session).generate(portfolio_id, tmp_path)
    workbook = load_workbook(path, data_only=True)
    exported = {
        row[0].value
        for row in workbook["Precios"].iter_rows(min_row=2)
        if row[0].value
    }
    workbook.close()
    assert exported == {"AMXL.MX"}
