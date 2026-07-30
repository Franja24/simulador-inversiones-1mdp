"""Pruebas unitarias e integración de Competition Intelligence."""

from io import BytesIO

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.orm import Session

from config.competition import CompetitionConfig
from config.quant_score import QuantScoreConfig
from config.simulation import MonteCarloConfig, PortfolioOptimizationConfig
from domain.competition import CompetitionDashboard
from repositories.competition_repository import CompetitionRepository
from services.competition_intelligence_service import CompetitionIntelligenceService
from services.competition_report_service import CompetitionReportService
from services.quant_score_service import QuantScoreService
from services.simulation_service import SimulationService
from tests.test_phase4_quant import seed_quant_history


def test_competition_configuration_requires_complete_normalized_weights() -> None:
    assert sum(CompetitionConfig().weights.values()) == pytest.approx(1)
    with pytest.raises(ValidationError):
        CompetitionConfig(weights={"aqs": 1})
    with pytest.raises(ValidationError):
        CompetitionConfig(
            weights={
                "monte_carlo": 1,
                "aqs": 1,
                "momentum": 0,
                "beating_benchmark": 0,
                "liquidity": 0,
                "risk": 0,
            }
        )


def test_liquidity_score_rewards_tight_spread_volume_and_execution() -> None:
    service = CompetitionIntelligenceService
    config = CompetitionConfig()
    liquid = service.liquidity_score("A", 0.002, 5_000_000, 100, config)
    illiquid = service.liquidity_score("B", 0.05, 100, 20, config)
    assert 0 <= illiquid.score < liquid.score <= 100
    assert liquid.components["spread"] > illiquid.components["spread"]
    assert liquid.execution_ease > illiquid.execution_ease


def test_competition_score_is_configurable_and_risk_adjusted() -> None:
    config = CompetitionConfig()
    safe, safe_parts = CompetitionIntelligenceService.competition_score(
        monte_carlo=80,
        aqs=80,
        momentum=80,
        beating_benchmark=0.8,
        liquidity=80,
        risk_penalty=10,
        config=config,
    )
    risky, _ = CompetitionIntelligenceService.competition_score(
        monte_carlo=80,
        aqs=80,
        momentum=80,
        beating_benchmark=0.8,
        liquidity=80,
        risk_penalty=90,
        config=config,
    )
    assert safe > risky
    assert safe == pytest.approx(sum(safe_parts.values()))


def test_rebalance_advisor_compares_net_benefit_cost_and_turnover() -> None:
    config = CompetitionConfig(
        minimum_rebalance_benefit_mxn=100,
        maximum_recommended_turnover=0.5,
    )
    advice = CompetitionIntelligenceService.rebalance_advice(
        {"A": 0.5, "B": 0.5},
        {"A": 0.4, "B": 0.6},
        1_000_000,
        0.01,
        0.03,
        config,
    )
    assert advice.recommend
    assert advice.expected_benefit > advice.expected_cost
    assert advice.turnover == pytest.approx(0.1)
    rejected = CompetitionIntelligenceService.rebalance_advice(
        {"A": 1},
        {"B": 1},
        1_000_000,
        0.03,
        0.02,
        config,
    )
    assert not rejected.recommend
    assert rejected.recommendation == "MANTENER"


def _competition_dashboard(
    session: Session, portfolio_id: int
) -> CompetitionDashboard:
    data = seed_quant_history(session, 100)
    effective = data["ALFA.MX"][-1].date
    symbols = ["ALFA.MX", "BETA.MX", "GAMA.MX"]
    QuantScoreService(session).calculate_universe(
        symbols,
        effective,
        "^MXX",
        QuantScoreConfig(minimum_history_rows=40),
        force=True,
    )
    simulation = MonteCarloConfig(
        horizons=[5],
        simulation_count=200,
        lookback_sessions=80,
        minimum_history_rows=40,
        sample_path_count=0,
        random_seed=23,
    )
    facade = SimulationService(session)
    for symbol in symbols:
        facade.simulate_symbol(symbol, effective, "^MXX", simulation)
    facade.optimize_portfolio(
        symbols,
        effective,
        "^MXX",
        simulation,
        PortfolioOptimizationConfig(
            candidate_count=30,
            minimum_symbols=2,
            maximum_symbols=3,
            minimum_confidence=0,
            random_seed=23,
        ),
    )
    return CompetitionIntelligenceService(session).build_dashboard(
        portfolio_id, effective, CompetitionConfig()
    )


def test_competition_integration_consumes_snapshots_persists_and_explains(
    session: Session, portfolio_id: int
) -> None:
    dashboard = _competition_dashboard(session, portfolio_id)
    assert 1 <= len(dashboard.top_candidates) <= 10
    assert dashboard.top_candidates[0].rank == 1
    assert dashboard.top_candidates[0].explanation.keys() >= {
        "AQS",
        "Monte Carlo",
        "Riesgo",
        "Régimen",
        "Liquidez",
        "Score final",
    }
    loaded = CompetitionRepository(session).load_snapshot(
        portfolio_id, dashboard.effective_date
    )
    assert loaded == dashboard


def test_daily_brief_and_all_report_exports(
    session: Session, portfolio_id: int
) -> None:
    dashboard = _competition_dashboard(session, portfolio_id)
    intelligence = CompetitionIntelligenceService(session)
    brief = intelligence.daily_brief(dashboard)
    assert "Régimen" in brief.markdown
    assert "Top candidatos" in brief.markdown
    assert brief.justification
    report = CompetitionReportService()
    assert report.markdown(brief).startswith(b"# Daily Brief")
    assert b"competition_score" in report.candidates_csv(dashboard)
    assert b"current_weight" in report.rebalance_csv(dashboard)
    assert report.pdf(brief).startswith(b"%PDF")
    workbook = load_workbook(BytesIO(report.excel(dashboard, brief)))
    assert workbook.sheetnames == [
        "Competition Dashboard",
        "Top Candidates",
        "Rebalance Advisor",
        "Daily Brief",
    ]
    assert workbook["Competition Dashboard"]["A1"].value == "Métrica"
