"""Exportaciones auditables de Competition Intelligence."""

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from domain.competition import CompetitionDashboard, DailyBrief


class CompetitionReportService:
    HEADER_FILL = PatternFill("solid", fgColor="17365D")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    @classmethod
    def excel(
        cls, dashboard: CompetitionDashboard, brief: DailyBrief
    ) -> bytes:
        workbook = Workbook()
        dashboard_sheet = workbook.active
        dashboard_sheet.title = "Competition Dashboard"
        cls._write_rows(
            dashboard_sheet,
            ["Métrica", "Valor"],
            [
                ["Fecha", dashboard.effective_date],
                ["Capital inicial", dashboard.capital_initial],
                ["Valor portafolio", dashboard.portfolio_value],
                ["Poder de compra", dashboard.buying_power],
                ["Benchmark", dashboard.benchmark_symbol],
                ["Rendimiento benchmark", dashboard.benchmark_return],
                ["Rendimiento portafolio", dashboard.portfolio_return],
                ["Diferencia", dashboard.excess_return],
                ["Régimen", dashboard.market_regime],
                ["Confianza", dashboard.confidence / 100],
                ["Riesgo", dashboard.risk_level],
                ["Estado modelo", dashboard.model_status],
                ["Última actualización", dashboard.last_update],
            ],
        )
        dashboard_sheet["B2"].number_format = "yyyy-mm-dd"
        for row in [3, 4, 5]:
            dashboard_sheet.cell(row=row, column=2).number_format = "$#,##0.00"
        for row in [7, 8, 9, 11]:
            dashboard_sheet.cell(row=row, column=2).number_format = "0.00%"

        candidates = workbook.create_sheet("Top Candidates")
        cls._write_rows(
            candidates,
            [
                "Rank",
                "Símbolo",
                "Competition Score",
                "AQS",
                "Monte Carlo",
                "Momentum",
                "Superar benchmark",
                "Liquidez",
                "Penalización riesgo",
                "Confianza",
                "Motivo principal",
            ],
            [
                [
                    item.rank,
                    item.symbol,
                    item.competition_score,
                    item.aqs,
                    item.monte_carlo,
                    item.momentum,
                    item.probability_beating_benchmark,
                    item.liquidity.score,
                    item.risk_penalty,
                    item.confidence,
                    item.main_reason,
                ]
                for item in dashboard.top_candidates
            ],
        )
        candidates.freeze_panes = "A2"
        candidates.auto_filter.ref = candidates.dimensions
        for cell in candidates["G"][1:]:
            cell.number_format = "0.0%"

        rebalance = workbook.create_sheet("Rebalance Advisor")
        cls._write_rows(
            rebalance,
            ["Métrica", "Valor"],
            [
                ["Recomendación", dashboard.rebalance.recommendation],
                ["Conviene", dashboard.rebalance.recommend],
                ["Costo esperado", dashboard.rebalance.expected_cost],
                ["Beneficio esperado", dashboard.rebalance.expected_benefit],
                ["Turnover", dashboard.rebalance.turnover],
                ["Justificación", dashboard.rebalance.justification],
                ["Compras", str(dashboard.rebalance.purchases)],
                ["Ventas", str(dashboard.rebalance.sales)],
            ],
        )
        rebalance["B4"].number_format = "$#,##0.00"
        rebalance["B5"].number_format = "$#,##0.00"
        rebalance["B6"].number_format = "0.0%"

        daily = workbook.create_sheet("Daily Brief")
        daily.append(["Daily Brief"])
        daily["A1"].fill = cls.HEADER_FILL
        daily["A1"].font = cls.HEADER_FONT
        for line in brief.markdown.splitlines():
            daily.append([line])
        daily.column_dimensions["A"].width = 110
        for row in daily.iter_rows():
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def dashboard_csv(dashboard: CompetitionDashboard) -> bytes:
        rows = [
            {"metric": key, "value": value}
            for key, value in dashboard.model_dump(mode="json").items()
            if key not in {"top_candidates", "rebalance"}
        ]
        return str(pd.DataFrame(rows).to_csv(index=False)).encode()

    @staticmethod
    def candidates_csv(dashboard: CompetitionDashboard) -> bytes:
        return str(
            pd.DataFrame(
                [item.model_dump(mode="json") for item in dashboard.top_candidates]
            ).to_csv(index=False)
        ).encode()

    @staticmethod
    def rebalance_csv(dashboard: CompetitionDashboard) -> bytes:
        changes = [
            {
                "symbol": symbol,
                "purchase": dashboard.rebalance.purchases.get(symbol, 0),
                "sale": dashboard.rebalance.sales.get(symbol, 0),
                "current_weight": dashboard.rebalance.current_weights.get(symbol, 0),
                "optimal_weight": dashboard.rebalance.optimal_weights.get(symbol, 0),
            }
            for symbol in sorted(
                set(dashboard.rebalance.current_weights)
                | set(dashboard.rebalance.optimal_weights)
            )
        ]
        return str(pd.DataFrame(changes).to_csv(index=False)).encode()

    @staticmethod
    def markdown(brief: DailyBrief) -> bytes:
        return brief.markdown.encode("utf-8")

    @staticmethod
    def pdf(brief: DailyBrief) -> bytes:
        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=letter,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
            title=f"Daily Brief {brief.effective_date}",
        )
        styles = getSampleStyleSheet()
        story = [
            Paragraph(
                f"Daily Brief - {brief.effective_date.isoformat()}",
                styles["Title"],
            ),
            Spacer(1, 8),
            Table(
                [
                    ["Mercado", brief.market],
                    ["Régimen", brief.regime],
                    ["Confianza", f"{brief.confidence:.1f}/100"],
                    ["Riesgo", brief.risk],
                    [
                        "Rebalanceo",
                        "Sí" if brief.rebalance_recommended else "No",
                    ],
                ],
                colWidths=[42 * mm, 120 * mm],
                style=TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#17365D")),
                        ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("PADDING", (0, 0), (-1, -1), 6),
                    ]
                ),
            ),
            Spacer(1, 12),
            Paragraph("Top candidatos", styles["Heading2"]),
            Paragraph(", ".join(brief.top_candidates) or "Sin candidatos", styles["BodyText"]),
            Spacer(1, 8),
            Paragraph("Recomendación", styles["Heading2"]),
            Paragraph(brief.recommendation, styles["BodyText"]),
            Paragraph(brief.justification, styles["BodyText"]),
            Spacer(1, 12),
            Paragraph(
                "Documento informativo; no constituye asesoría financiera.",
                styles["Italic"],
            ),
        ]
        document.build(story)
        return output.getvalue()

    @classmethod
    def _write_rows(
        cls,
        sheet: object,
        headers: list[str],
        rows: list[list[object]],
    ) -> None:
        sheet.append(headers)  # type: ignore[attr-defined]
        for cell in sheet[1]:  # type: ignore[index]
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
        for row in rows:
            sheet.append(row)  # type: ignore[attr-defined]
        sheet.freeze_panes = "A2"  # type: ignore[attr-defined]
        for column in sheet.columns:  # type: ignore[attr-defined]
            width = min(60, max(len(str(cell.value or "")) for cell in column) + 2)
            sheet.column_dimensions[column[0].column_letter].width = width  # type: ignore[attr-defined]
        for row in sheet.iter_rows():  # type: ignore[attr-defined]
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
