"""Generación de reportes Excel de la Fase 2."""

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from config.settings import Settings, get_settings
from repositories.price_repository import SqlPriceRepository
from repositories.transaction_repository import TransactionRepository
from services.portfolio_service import PortfolioService


class ReportService:
    """Construye un libro Excel legible incluso cuando no hay operaciones."""

    SHEETS = ["Resumen", "Posiciones", "Operaciones", "Precios", "Reglas"]

    def __init__(self, session: Session, settings: Settings | None = None) -> None:
        self.session = session
        self.settings = settings or get_settings()

    def generate(self, portfolio_id: int, output_dir: Path | None = None) -> Path:
        """Genera, da formato y guarda un reporte completo."""
        portfolio_service = PortfolioService(self.session)
        portfolio = portfolio_service.get_required(portfolio_id)
        prices = SqlPriceRepository(self.session)
        valuation = portfolio_service.valuation(portfolio_id, prices)
        positions = portfolio_service.calculate_positions(portfolio_id, prices)
        transactions = TransactionRepository(self.session).list_for_portfolio(portfolio_id)
        price_history = prices.list_all()

        workbook = Workbook()
        workbook.remove(workbook.active)
        summary = workbook.create_sheet("Resumen")
        summary.append(["Concepto", "Valor"])
        for label, value in [
            ("Nombre del portafolio", portfolio.name),
            ("Capital inicial", portfolio.initial_capital),
            ("Valor actual", valuation["total"]),
            ("Efectivo", valuation["cash"]),
            ("Invertido", valuation["invested"]),
            ("Ganancia total", valuation["profit_loss"]),
            ("Ganancia realizada", valuation["realized"]),
            ("Ganancia no realizada", valuation["unrealized"]),
            ("Rendimiento", valuation["return_percentage"] / Decimal("100")),
            ("Fecha de generación", datetime.now()),
        ]:
            summary.append([label, value])

        self._table(
            workbook.create_sheet("Posiciones"),
            [
                "Emisora",
                "Empresa",
                "Cantidad",
                "Precio promedio",
                "Precio actual",
                "Invertido",
                "Valor de mercado",
                "Ganancia/Pérdida",
                "Rendimiento",
                "Peso",
                "Stop loss",
                "Take profit",
                "Fecha último precio",
            ],
            [
                [
                    item.symbol,
                    item.company_name,
                    item.total_quantity,
                    item.average_purchase_price,
                    item.current_price,
                    item.invested_amount,
                    item.current_market_value,
                    item.unrealized_profit_loss,
                    item.unrealized_return_percentage / Decimal("100"),
                    item.portfolio_weight / Decimal("100"),
                    item.stop_loss,
                    item.take_profit,
                    item.last_price_date,
                ]
                for item in positions
            ],
        )
        self._table(
            workbook.create_sheet("Operaciones"),
            [
                "Fecha",
                "Tipo",
                "Emisora",
                "Empresa",
                "Cantidad",
                "Precio",
                "Comisión",
                "Impuestos",
                "Total",
                "Estrategia",
                "Motivo",
                "Stop loss",
                "Take profit",
                "Notas",
            ],
            [
                [
                    item.transaction_date,
                    item.transaction_type.value,
                    item.symbol,
                    item.company_name,
                    item.quantity,
                    item.price,
                    item.commission,
                    item.taxes,
                    item.total_amount,
                    item.strategy,
                    item.reason,
                    item.stop_loss,
                    item.take_profit,
                    item.notes,
                ]
                for item in transactions
            ],
        )
        self._table(
            workbook.create_sheet("Precios"),
            ["Emisora", "Precio", "Fecha", "Proveedor", "Notas"],
            [
                [item.symbol, item.price, item.price_date, item.provider, item.notes]
                for item in price_history
            ],
        )
        alerts = [
            f"{item.symbol}: peso {item.portfolio_weight:.2f}%"
            for item in positions
            if item.portfolio_weight
            > Decimal(str(self.settings.max_position_weight * 100))
        ]
        self._table(
            workbook.create_sheet("Reglas"),
            ["Regla", "Valor"],
            [
                ["Peso máximo", self.settings.max_position_weight],
                ["Mínimo de emisoras", self.settings.min_different_symbols],
                ["Alertas actuales", "; ".join(alerts) if alerts else "Sin alertas"],
            ],
        )
        self._format(workbook)
        destination = output_dir or Path("data/exports")
        destination.mkdir(parents=True, exist_ok=True)
        filename = f"reporte_reto_actinver_{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx"
        path = destination / filename
        workbook.save(path)
        load_workbook(path).close()
        return path

    @staticmethod
    def _table(sheet, headers: list[str], rows: list[list[object]]) -> None:  # type: ignore[no-untyped-def]
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    value.replace(tzinfo=None)
                    if isinstance(value, datetime) and value.tzinfo
                    else value
                    for value in row
                ]
            )

    @staticmethod
    def _format(workbook: Workbook) -> None:
        currency_words = {
            "capital",
            "valor",
            "efectivo",
            "invertido",
            "ganancia",
            "precio",
            "comisión",
            "impuestos",
            "total",
            "stop",
            "take",
        }
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            for column in sheet.columns:
                letter = get_column_letter(column[0].column)
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 40)
                sheet.column_dimensions[letter].width = max(width, 12)
                header = str(column[0].value or "").lower()
                if any(word in header for word in currency_words):
                    for cell in column[1:]:
                        if isinstance(cell.value, (int, float, Decimal)):
                            cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
                if "rendimiento" in header or "peso" in header:
                    for cell in column[1:]:
                        cell.number_format = "0.00%"
